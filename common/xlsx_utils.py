"""
xlsx_utils.py — shared openpyxl utilities for the data-cleaning eval.

Used by:
  - programmatic_checks.py  (formula inspection, cell iteration)
  - llm_judge.py            (extract_to_text — converts workbook to judge-readable string)
"""

from __future__ import annotations

from pathlib import Path
from typing import Iterator

import openpyxl
from openpyxl.cell.cell import Cell
from openpyxl.worksheet.worksheet import Worksheet

# ── Constants ─────────────────────────────────────────────────────────────────

# Frozen set for O(1) membership test — used in hot loop over all cells.
# openpyxl returns these as literal string values when data_only=False AND the
# cached value (from the last Excel save) was an error.
EXCEL_ERRORS: frozenset[str] = frozenset({
    "#REF!", "#VALUE!", "#DIV/0!", "#NAME?", "#NULL!", "#N/A", "#NUM!",
})


# ── Workbook loading ──────────────────────────────────────────────────────────

def load_workbook(path: Path | str, data_only: bool = False):
    """
    Thin wrapper around openpyxl.load_workbook.

    CRITICAL distinction:
      data_only=False (default):
        Cell.value is the *formula string* for formula cells, e.g. '=SUM(A1:A10)'.
        Use this when you need to inspect formula structure.

      data_only=True:
        Cell.value is the *cached calculated result* (the number Excel last computed).
        WARNING: if you load with data_only=True and then save, formulas are
        permanently replaced with their values — the workbook is corrupted for
        formula checking. Never resave after data_only=True.
    """
    return openpyxl.load_workbook(str(path), data_only=data_only)


# ── Cell iterators ────────────────────────────────────────────────────────────

def iter_formula_cells(ws: Worksheet) -> Iterator[Cell]:
    """
    Yield every cell in a worksheet whose value is an Excel formula.

    Detection: cell.value is a string starting with '='.
    This works because openpyxl stores formulas as their source strings
    when loaded with data_only=False.

    Example yielded values: '=SUM(A1:A10)', "=SUMIF('Cleaned Data'!D$2:D$71,\"North\",...)"
    """
    for row in ws.iter_rows():
        for cell in row:
            if isinstance(cell.value, str) and cell.value.startswith("="):
                yield cell


def iter_error_cells(ws: Worksheet) -> Iterator[Cell]:
    """
    Yield every cell containing an Excel error value.

    openpyxl reports formula errors as the error string itself (e.g. '#REF!'),
    not as a special error type — hence the string membership check.

    Note: this only works reliably when the .xlsx was saved by Excel or LibreOffice
    after recalculation. Files created purely by openpyxl and not recalculated
    will have formula strings, not cached error values. Use recalc.py first.
    """
    for row in ws.iter_rows():
        for cell in row:
            if isinstance(cell.value, str) and cell.value in EXCEL_ERRORS:
                yield cell


def iter_all_cells(wb) -> Iterator[tuple[str, Cell]]:
    """Yield (sheet_name, cell) for every cell in every worksheet."""
    for ws in wb.worksheets:
        for row in ws.iter_rows():
            for cell in row:
                yield ws.title, cell


# ── Row / column helpers ──────────────────────────────────────────────────────

def count_data_rows(ws: Worksheet, id_col: int = 1) -> int:
    """
    Count non-empty data rows, excluding the header (row 1).

    Uses id_col (1-indexed, default=1 = order_id) as the presence indicator.
    A row is counted if that column cell is non-None and non-empty-string.

    Why not ws.max_row?
      openpyxl's max_row includes rows that were once populated and then cleared.
      It tracks the highest *ever-written* row, not the current populated row count.
      Iterating and checking the key column is more reliable for sparse sheets.
    """
    count = 0
    for row in ws.iter_rows(min_row=2, min_col=id_col, max_col=id_col, values_only=True):
        val = row[0]
        if val is not None and str(val).strip() != "":
            count += 1
    return count


def get_column_cells(ws: Worksheet, col_idx: int, start_row: int = 2) -> list[Cell]:
    """
    Return all Cell objects in a column (1-indexed) from start_row onwards.
    Includes empty cells — caller decides how to handle None values.
    """
    cells = []
    for row in ws.iter_rows(min_row=start_row, min_col=col_idx, max_col=col_idx):
        cells.append(row[0])
    return cells


# ── LLM judge text extraction ─────────────────────────────────────────────────

def extract_to_text(wb) -> str:
    """
    Convert a workbook to a text summary for LLM judge consumption (Day 17).

    Captures per sheet:
      - Column headers
      - Approximate row count
      - First 3 data rows as samples
      - Up to 3 formula examples (signals to the judge whether formulas are present)

    The judge sees this text, not the raw .xlsx binary, so the extraction must
    convey both structure (headers, layout) and substance (sample values, formula types).
    """
    sections: list[str] = []

    for sheet_name in wb.sheetnames:
        ws = wb[sheet_name]
        lines: list[str] = [f"=== Sheet: {sheet_name} ==="]
        all_rows = list(ws.iter_rows(values_only=False))

        if not all_rows:
            lines.append("  [empty]")
            sections.append("\n".join(lines))
            continue

        # Headers (row 1)
        headers = [str(c.value) if c.value is not None else "" for c in all_rows[0]]
        lines.append(f"  Columns ({len(headers)}): {', '.join(h for h in headers if h)}")
        lines.append(f"  Data rows (approx): {max(0, ws.max_row - 1)}")

        # First 3 data rows
        for row in all_rows[1:4]:
            vals = [str(c.value) if c.value is not None else "" for c in row]
            lines.append(f"  Sample: {', '.join(vals)}")

        # Formula examples (up to 3) — tells judge whether model used formulas
        examples: list[str] = []
        for row in all_rows[1:]:
            for cell in row:
                if isinstance(cell.value, str) and cell.value.startswith("="):
                    examples.append(f"    {cell.coordinate}: {cell.value}")
                    if len(examples) >= 3:
                        break
            if len(examples) >= 3:
                break

        if examples:
            lines.append("  Formula examples:")
            lines.extend(examples)
        else:
            lines.append("  Formula examples: [none — all values may be hardcoded]")

        sections.append("\n".join(lines))

    return "\n\n".join(sections)
